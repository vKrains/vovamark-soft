#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from datetime import datetime
from zoneinfo import ZoneInfo
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# === Настройки ===
FILES = [
    "D:/Софт/скрипты и аутпутс/на закупку/НА_ЗАКУПКУ_КРАСНОДАР.xlsx",
    "D:/Софт/скрипты и аутпутс/на закупку/НА_ЗАКУПКУ_КАЛЕДИНО.xlsx",
    "D:/Софт/скрипты и аутпутс/на закупку/НА_ЗАКУПКУ_МОСКВА.xlsx"
]
DATE_HEADER = "Дата"   # <-- имя колонки с датой (как в первой строке!)
MOSCOW_TZ = ZoneInfo("Europe/Moscow")

# Цвета (HEX без #)
YELLOW = "FFF59D"    # 24–36 ч
ORANGE = "FFCC80"    # 36–50 ч
LIGHT_RED = "FFCDD2" # >50 ч


def parse_dt(value):
    """Парсим дату из ячейки Excel, возвращаем datetime с TZ Europe/Moscow или None"""
    if value is None:
        return None

    if isinstance(value, datetime):
        dt = value
    else:
        s = str(value).strip()
        if not s:
            return None
        try:
            dt = datetime.fromisoformat(s)
        except Exception:
            try:
                dt = datetime.strptime(s, "%Y-%m-%d %H:%M:%S")
            except Exception:
                return None

    if dt.tzinfo is None:
        return dt.replace(tzinfo=MOSCOW_TZ)
    else:
        return dt.astimezone(MOSCOW_TZ)


def get_header_index(ws, header_name):
    """Возвращает номер колонки по имени заголовка (строка 1)"""
    for col_idx in range(1, ws.max_column + 1):
        if str(ws.cell(row=1, column=col_idx).value).strip() == header_name:
            return col_idx
    return None


def fill_for_hours(hours):
    if hours > 30:
        return PatternFill(start_color=LIGHT_RED, end_color=LIGHT_RED, fill_type="solid")
    elif hours >= 20:
        return PatternFill(start_color=ORANGE, end_color=ORANGE, fill_type="solid")
    return None


def process_file(path):
    print(f"Обработка: {path}")
    wb = load_workbook(path)
    ws = wb.active

    col_idx = get_header_index(ws, DATE_HEADER)
    if not col_idx:
        print(f"  ❌ Не найдена колонка '{DATE_HEADER}' в файле {path}")
        return

    now = datetime.now(MOSCOW_TZ)
    colored = 0

    for row in range(2, ws.max_row + 1):
        cell_val = ws.cell(row=row, column=col_idx).value
        dt = parse_dt(cell_val)
        if not dt:
            continue
        hours = (now - dt).total_seconds() / 3600.0
        fill = fill_for_hours(hours)
        if fill:
            for c in range(1, ws.max_column + 1):
                ws.cell(row=row, column=c).fill = fill
            colored += 1

    wb.save(path)
    print(f"  ✅ Готово. Подсвечено строк: {colored}")


def main():
    for file in FILES:
        try:
            process_file(file)
        except Exception as e:
            print(f"Ошибка при обработке {file}: {e}")


if __name__ == "__main__":
    main()
