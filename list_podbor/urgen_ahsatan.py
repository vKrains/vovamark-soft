import os
import re
import random
from datetime import datetime, timedelta

from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Border, Side, Font, PatternFill, Alignment


FOLDER = r"D:/Софт/скрипты и аутпутс/Листы подбора/обработка"

DATE_START = datetime(2026, 5, 31)
DATE_END   = datetime(2027, 1, 31)
DELTA_DAYS = (DATE_END - DATE_START).days


def extract_wb(value):
    if value is None:
        return ""
    m = re.search(r"WB-[A-Za-z0-9-]+", str(value))
    return m.group(0) if m else ""


def process_file(path):
    wb = load_workbook(path)
    ws = wb.active

    # значение поставки (берём ДО удаления строк)
    supply_value = extract_wb(ws["A2"].value)

    # 1️⃣ снять объединения в строках 1–4
    merges_to_remove = []
    for merged in ws.merged_cells.ranges:
        if merged.min_row <= 4:
            merges_to_remove.append(merged)

    for merged in merges_to_remove:
        ws.unmerge_cells(str(merged))

    # 2️⃣ УДАЛИТЬ строки 1–4 физически
    ws.delete_rows(1, 4)

    # теперь заголовок стал первой строкой
    header_row = 1
    max_col = ws.max_column
    max_row = ws.max_row

    # взять стиль существующего заголовка
    tmpl = ws.cell(row=header_row, column=1)

    header_font = Font(
        name=tmpl.font.name,
        size=tmpl.font.size,
        bold=tmpl.font.bold,
        color=tmpl.font.color
    )

    header_fill = PatternFill(
        fill_type=tmpl.fill.fill_type,
        fgColor=getattr(tmpl.fill, "fgColor", None)
    )

    header_align = Alignment(
        horizontal=tmpl.alignment.horizontal,
        vertical=tmpl.alignment.vertical,
        wrap_text=tmpl.alignment.wrap_text
    )

    # 3️⃣ добавить новые заголовки
    ws.cell(row=header_row, column=max_col + 1, value="Поставка")
    ws.cell(row=header_row, column=max_col + 2, value="Собрано")
    ws.cell(row=header_row, column=max_col + 3, value="Собрал")
    ws.cell(row=header_row, column=max_col + 4, value="Срок годности")

    for c in range(max_col + 1, max_col + 5):
        cell = ws.cell(row=header_row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # границы
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for r in range(header_row, max_row + 1):
        for c in range(max_col + 1, max_col + 5):
            ws.cell(row=r, column=c).border = border

    # 4️⃣ заполнить колонку "Поставка"
    for r in range(header_row + 1, max_row + 1):
        ws.cell(row=r, column=max_col + 1, value=supply_value)

    # 5️⃣ выпадающие списки
    dv_yes_no = DataValidation(
        type="list",
        formula1='"Да,Нет"',
        allow_blank=True
    )

    dv_workers = DataValidation(
        type="list",
        formula1='"Сборщик1,Сборщик2,Сборщик3"',
        allow_blank=True
    )

    ws.add_data_validation(dv_yes_no)
    ws.add_data_validation(dv_workers)

    dv_yes_no.add(
        f"{ws.cell(header_row+1, max_col+2).coordinate}:"
        f"{ws.cell(max_row, max_col+2).coordinate}"
    )

    dv_workers.add(
        f"{ws.cell(header_row+1, max_col+3).coordinate}:"
        f"{ws.cell(max_row, max_col+3).coordinate}"
    )

    # 6️⃣ случайные даты срока годности
    for r in range(header_row + 1, max_row + 1):
        d = DATE_START + timedelta(days=random.randint(0, DELTA_DAYS))
        cell = ws.cell(row=r, column=max_col + 4, value=d)
        cell.number_format = "DD.MM.YYYY"

    wb.save(path)
    print(f"Обновлён: {path}")


def main():
    for file in os.listdir(FOLDER):
        if file.lower().endswith(".xlsx") and not file.startswith("~$"):
            process_file(os.path.join(FOLDER, file))


if __name__ == "__main__":
    main()
